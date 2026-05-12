"""
Excel Report Generator for UV Oil Leak Detection
Columns: VIN Number | Timestamp | Status | Detection Result | Image Feedback (embedded snapshot)
"""

import os
from datetime import datetime
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
import io

REPORTS_DIR = "reports"


class ExcelReporter:
    def __init__(self):
        os.makedirs(REPORTS_DIR, exist_ok=True)
        self.today_str = datetime.now().strftime("%Y%m%d")
        self.report_path = os.path.join(REPORTS_DIR, f"OilLeak_Report_{self.today_str}.xlsx")
        self._records = []

    # ------------------------------------------------------------------ #
    #  Public API
    # ------------------------------------------------------------------ #

    def add_record(self, win_number: str, status: str, image_path: Optional[str]):
        """Add one inspection record."""
        self._records.append({
            "win": win_number,
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "status": status,               # "OK" or "NOT OK"
            "image_path": image_path or "",
        })

    def generate(self) -> str:
        """Write/append report to disk. Returns the file path."""
        if os.path.exists(self.report_path):
            wb = load_workbook(self.report_path)
        else:
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        # Detail sheet for this session
        sheet_name = f"Session_{datetime.now().strftime('%H%M%S')}"
        ws_detail = wb.create_sheet(title=sheet_name)
        self._write_detail_sheet(ws_detail)

        # Always regenerate summary as first sheet
        if "Summary" in wb.sheetnames:
            del wb["Summary"]
        ws_summary = wb.create_sheet(title="Summary", index=0)
        self._write_summary_sheet(ws_summary)

        wb.save(self.report_path)
        return self.report_path

    # ------------------------------------------------------------------ #
    #  Styling helpers
    # ------------------------------------------------------------------ #

    def _border(self):
        s = Side(border_style="thin", color="2A3A4A")
        return Border(left=s, right=s, top=s, bottom=s)

    def _cell(self, ws, row, col, value,
              bg="0D1525", fg="CCDDEE", bold=False,
              size=10, align="center", wrap=False):
        c = ws.cell(row=row, column=col, value=value)
        c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(bold=bold, color=fg, size=size, name="Calibri")
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        c.border = self._border()
        return c

    # ------------------------------------------------------------------ #
    #  Detail sheet
    # ------------------------------------------------------------------ #

    def _write_detail_sheet(self, ws):
        # --- Title row ---
        ws.merge_cells("A1:E1")
        c = ws["A1"]
        c.value = "⚙  UV FLUORESCENT ENGINE OIL LEAK DETECTION REPORT"
        c.font = Font(bold=True, size=15, color="00D4FF", name="Calibri")
        c.fill = PatternFill("solid", fgColor="060A18")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 36

        # --- Sub-header ---
        ws.merge_cells("A2:E2")
        c2 = ws["A2"]
        c2.value = (
            f"Generated: {datetime.now().strftime('%Y-%m-%d  %H:%M:%S')}    "
            f"|    Total Records: {len(self._records)}"
        )
        c2.font = Font(italic=True, size=9, color="778899")
        c2.fill = PatternFill("solid", fgColor="0A0F1E")
        c2.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 18

        # --- Column headers ---
        headers = ["#", "VIN / WIN Number", "Timestamp", "Detection Result", "Image Feedback (High Quality)"]
        widths  = [5, 22, 22, 28, 68]   # column E widened for high-res image

        for col_i, (h, w) in enumerate(zip(headers, widths), start=1):
            ltr = get_column_letter(col_i)
            ws.column_dimensions[ltr].width = w
            c = ws.cell(row=3, column=col_i, value=h)
            c.font  = Font(bold=True, color="00D4FF", size=11, name="Calibri")
            c.fill  = PatternFill("solid", fgColor="0A1A2E")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = self._border()
        ws.row_dimensions[3].height = 24

        # --- Data rows ---
        for i, rec in enumerate(self._records, start=1):
            row = i + 3
            is_ok = rec["status"] == "OK"

            # Alternating row background
            row_bg  = "081508" if is_ok else "180808"   # dark green / dark red tint
            txt_fg  = "CCEECC" if is_ok else "EEBBBB"

            # # column
            self._cell(ws, row, 1, i, bg=row_bg, fg="889999")

            # VIN
            self._cell(ws, row, 2, rec["win"], bg=row_bg, fg="DDEEFF", bold=True, align="left")

            # Timestamp
            self._cell(ws, row, 3, rec["timestamp"], bg=row_bg, fg="AABBCC")

            # Detection Result — colour coded
            result_text = "✅  DETECTED: NO OIL LEAK" if is_ok else "❌  DETECTED: OIL LEAK PRESENT"
            result_fg   = "00FF88" if is_ok else "FF4444"
            self._cell(ws, row, 4, result_text, bg=row_bg, fg=result_fg, bold=True)

            # Image Feedback — embed HIGH QUALITY image in column E
            img_path = rec.get("image_path", "")
            if img_path and os.path.exists(img_path):
                try:
                    pil = PILImage.open(img_path)
                    # High-quality: scale down only if larger than 480×300, keep aspect ratio
                    MAX_W, MAX_H = 480, 300
                    pil.thumbnail((MAX_W, MAX_H), PILImage.LANCZOS)
                    buf = io.BytesIO()
                    pil.save(buf, format="PNG", optimize=False, compress_level=1)
                    buf.seek(0)
                    xl_img = XLImage(buf)
                    xl_img.width  = pil.width
                    xl_img.height = pil.height
                    ws.add_image(xl_img, f"E{row}")
                    # Row height ~ image height in pts (1 px ≈ 0.75 pt)
                    ws.row_dimensions[row].height = pil.height * 0.75 + 4
                    # Filename label behind image
                    fc = ws.cell(row=row, column=5, value=os.path.basename(img_path))
                    fc.fill = PatternFill("solid", fgColor=row_bg)
                    fc.font = Font(color="334455", size=8)
                except Exception as ex:
                    self._cell(ws, row, 5, f"[Image error: {ex}]", bg=row_bg, fg="FF8888")
                    ws.row_dimensions[row].height = 22
            else:
                self._cell(ws, row, 5,
                           "No leak image captured" if is_ok else "Image not found",
                           bg=row_bg, fg="556677")
                ws.row_dimensions[row].height = 22

        # Freeze header rows
        ws.freeze_panes = "A4"

    # ------------------------------------------------------------------ #
    #  Summary sheet
    # ------------------------------------------------------------------ #

    def _write_summary_sheet(self, ws):
        total  = len(self._records)
        passes = sum(1 for r in self._records if r["status"] == "OK")
        fails  = total - passes
        rate   = (passes / total * 100) if total > 0 else 0

        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 22

        # Title
        ws.merge_cells("A1:B1")
        t = ws["A1"]
        t.value = "INSPECTION SUMMARY"
        t.font  = Font(bold=True, size=18, color="00D4FF", name="Calibri")
        t.fill  = PatternFill("solid", fgColor="060A18")
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 44

        def kv(row, key, val, val_color="FFFFFF", bg_k="0A1220", bg_v="08101C"):
            k = ws.cell(row=row, column=1, value=key)
            k.font  = Font(bold=True, color="7799AA", size=11, name="Calibri")
            k.fill  = PatternFill("solid", fgColor=bg_k)
            k.alignment = Alignment(horizontal="left", vertical="center", indent=2)
            k.border = self._border()

            v = ws.cell(row=row, column=2, value=val)
            v.font  = Font(bold=True, color=val_color, size=14, name="Calibri")
            v.fill  = PatternFill("solid", fgColor=bg_v)
            v.alignment = Alignment(horizontal="center", vertical="center")
            v.border = self._border()
            ws.row_dimensions[row].height = 30

        kv(2, "Report Date",       datetime.now().strftime("%Y-%m-%d"), "AACCEE")
        kv(3, "Total VINs Inspected", total,   "00D4FF")
        kv(4, "✅  Passed  (No Leak)",  passes,  "00FF88")
        kv(5, "❌  Failed  (Leak Found)", fails, "FF4444")
        kv(6, "Pass Rate %",
           f"{rate:.1f}%",
           "00FF88" if rate >= 80 else "FFAA00" if rate >= 50 else "FF4444")
        kv(7, "Report Generated At", datetime.now().strftime("%H:%M:%S"), "AABBCC")

        ws.freeze_panes = "A2"
